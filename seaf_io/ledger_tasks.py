import logging
import os
import stat
from datetime import datetime

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from seaserv import seafile_api

from seafevents.seaf_io.utils import height_transfer, width_transfer, is_int_str, gen_decimal_format
from seafevents.app.config import DTABLE_WEB_SERVER, DTABLE_WEB_LEDGER_API_TOKEN, LEDGER_TABLE_NAME
from seafevents.utils.seatable_api import SeaTableAPI, ColumnTypes
from seafevents.seaf_io.utils import TaskError

logger = logging.getLogger(__name__)


def list_repo_dir_ledger_rows(seatable_api: SeaTableAPI, fields, repo_id, dir=None):
    dir = dir or '/'
    rows = []
    dirents = seafile_api.list_dir_by_path(repo_id, dir)
    files, sub_dirs = [], []
    for dirent in dirents:
        if stat.S_ISDIR(dirent.mode):
            sub_dirs.append(os.path.join(dir, dirent.obj_name))
        else:
            files.append(os.path.join(dir, dirent.obj_name))
    step = 10000
    fields_str = ', '.join(map(lambda x: f"`{x}`", fields))
    for i in range(0, len(files), step):
        files_str = ', '.join(map(lambda x: f"'{x}'", files[i: i+step]))
        sql = f"SELECT {fields_str} FROM `{LEDGER_TABLE_NAME}` WHERE `Repo ID`='{repo_id}' AND `文件路径` in ({files_str})"
        result = seatable_api.query(sql, convert=True)
        rows.extend(result['results'])
    for sub_dir in sub_dirs:
        sub_rows = list_repo_dir_ledger_rows(seatable_api, fields, repo_id, sub_dir)
        rows.extend(sub_rows)
    return rows


def genetate_ws_cell(ws, cell_value, column_type, column_data=None):
    if not cell_value and not isinstance(cell_value, int) and not isinstance(cell_value, float):
        c = WriteOnlyCell(ws, value=None)
    elif column_type == ColumnTypes.TEXT:
        c = WriteOnlyCell(ws, value=str(cell_value))
    elif column_type == ColumnTypes.NUMBER:
        # if value cannot convert to float or int, just pass
        if isinstance(cell_value, int) or isinstance(cell_value, float):
            c = WriteOnlyCell(ws, value=cell_value)
            c.number_format = gen_decimal_format(str(cell_value))
        else:
            try:
                cell_value = str(cell_value)
                if is_int_str(cell_value):
                    c = WriteOnlyCell(ws, value=int(cell_value))
                else:
                    c = WriteOnlyCell(ws, value=float(cell_value))
            except:
                c = WriteOnlyCell(ws, value=None)
            else:
                c.number_format = gen_decimal_format(cell_value)
    elif column_type == ColumnTypes.DATE:
        try:
            date_value = datetime.fromisoformat(cell_value)
        except Exception as e:
            logger.warning('convert date value %s to date error: %s', cell_value, e)
            c = WriteOnlyCell(ws, value=str(cell_value))
        else:
            c = WriteOnlyCell(ws, value=date_value.strftime('%Y-%m-%d %H:%M'))
    elif column_type == ColumnTypes.SINGLE_SELECT:
        c = WriteOnlyCell(ws, value=str(cell_value))
    elif column_type == ColumnTypes.MULTIPLE_SELECT:
        c = WriteOnlyCell(ws, value=', '.join(str(v) for v in cell_value))
    elif column_type == ColumnTypes.FORMULA:
        result_type = column_data.get('result_type')
        # number string date date bool array
        if result_type == 'number':
            c = genetate_ws_cell(ws, cell_value, ColumnTypes.NUMBER)
        elif result_type == 'string':
            c = genetate_ws_cell(ws, cell_value, ColumnTypes.TEXT)
        elif result_type == 'date':
            c = genetate_ws_cell(ws, cell_value, ColumnTypes.DATE)
        elif result_type == 'bool':
            ws_cell_value = None
            if cell_value:
                ws_cell_value = 'True'
            else:
                ws_cell_value = 'False'
            c = WriteOnlyCell(ws, value=ws_cell_value)
        elif result_type == 'array':
            pass
        else:
            c = WriteOnlyCell(ws, value=str(cell_value))
    else:
        c = WriteOnlyCell(ws, value=str(cell_value))
    return c


def generate_ws_row(ws, row, columns):
    """row -> openpyxl cells order by columns
    
    Keyword arguments:
    row -- {col_name: col_value}
    Return: a list of openpyxl cell
    """
    cells = []
    for col in columns:
        column_name = col['name']
        column_type = col['type']
        column_data = col.get('data') or {}
        cell_value = row.get(column_name)
        c = genetate_ws_cell(ws, cell_value, column_type, column_data=column_data)
        cells.append(c)
    return cells


def export_ledger_to_excel(repo_id, parent_dir=None):
    """export ledger table rows to a tmp excel to /tmp

    run successfully or raise some error
    """
    logger.info('NOW: %s', ((DTABLE_WEB_SERVER, DTABLE_WEB_LEDGER_API_TOKEN, LEDGER_TABLE_NAME),))
    if not all((DTABLE_WEB_SERVER, DTABLE_WEB_LEDGER_API_TOKEN, LEDGER_TABLE_NAME)):
        raise TaskError('Feature not enabled', 403)
    try:
        seatable_api = SeaTableAPI(DTABLE_WEB_LEDGER_API_TOKEN, DTABLE_WEB_SERVER)
    except Exception as e:
        raise TaskError('Internal Server Error', 500)
    try:
        metadata = seatable_api.get_metadata()
    except Exception as e:
        logger.error('ledger api token: %s metadata error: %s', DTABLE_WEB_LEDGER_API_TOKEN, e)
        raise TaskError('Internal Server Error', 500)
    ledger_table = None
    for table in metadata['tables']:
        if table['name'] == LEDGER_TABLE_NAME:
            ledger_table = table
            break
    if not ledger_table:
        raise TaskError('Ledger table: %s not found', 404)
    columns_name_dict = {col['name']: col for col in ledger_table['columns']}
    fields = ['Repo ID', '文件名', '文件路径', '文件大分类', '文件中分类', '文件小分类', '密级', '保密期限', '创建日期', '废弃日期']
    field_columns = []
    for field in fields:
        field_column = columns_name_dict.get(field)
        if not field_column:
            raise TaskError('Ledger table column: %s not found' % field, 404)
        field_columns.append(field_column)
    try:
        rows = list_repo_dir_ledger_rows(seatable_api, fields, repo_id, parent_dir)
    except Exception as e:
        logger.exception('list repo: %s parent_dir: %s error: %s', repo_id, parent_dir, e)
        raise TaskError('Internal Server Error', 500)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(LEDGER_TABLE_NAME)
    ws.row_dimensions[1].height = height_transfer('default')
    head_cell_list = []
    for index, field in enumerate(fields, start=1):
        c = WriteOnlyCell(ws, value=field)
        c_pos = get_column_letter(index)
        c_width_xls = width_transfer(200)
        ws.column_dimensions[c_pos].width = c_width_xls
        head_cell_list.append(c)
    ws.append(head_cell_list)
    for index, row in enumerate(rows, start=2):
        ws_cell_list = generate_ws_row(ws, row, field_columns)
        ws.row_dimensions[index].height = height_transfer('default')
        ws.append(ws_cell_list)
    target_dir = '/tmp/seafile-io/ledgers-excels/' + repo_id
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)
    target_path = os.path.join(target_dir, f"{DTABLE_WEB_LEDGER_API_TOKEN}-{LEDGER_TABLE_NAME}-{parent_dir.replace('/', '-')}.xlsx")
    wb.save(target_path)
