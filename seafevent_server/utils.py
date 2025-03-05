import os
import ast
import logging
import time

import openpyxl
import datetime
import pytz
import os
import json
import uuid
import posixpath
import random
import hashlib

from sqlalchemy import desc, select, text

from seafevents.events.models import FileAudit, FileUpdate, PermAudit, UserLogin
from seafevents.app.config import TIME_ZONE
from seafevents.utils.ccnet_db import CcnetDB
from seafevents.utils.seafile_db import SeafileDB
from seafevents.repo_metadata.constants import ZERO_OBJ_ID
from seafevents.repo_data import repo_data
from seafevents.utils.md2sdoc import md2sdoc
from seafevents.utils.constants import WIKI_PAGES_DIR, WIKI_CONFIG_PATH, \
    WIKI_CONFIG_FILE_NAME

from seaserv import get_org_id_by_repo_id, seafile_api, get_commit
from seafobj import CommitDiffer, commit_mgr, fs_mgr
from seafobj.exceptions import GetObjectError

logger = logging.getLogger('seafevents')

WIKI_FILE_TMP_DIR = '/tmp'
SYS_DIR_PATHS = ['images']


def get_diff_files(repo_id, old_commit_id, new_commit_id):
    if old_commit_id == new_commit_id:
        return

    old_root = None
    if old_commit_id:
        try:
            old_commit = commit_mgr.load_commit(repo_id, 0, old_commit_id)
            old_root = old_commit.root_id
        except GetObjectError as e:
            logger.debug(e)
            old_root = None

    try:
        new_commit = commit_mgr.load_commit(repo_id, 0, new_commit_id)
    except GetObjectError as e:
        # new commit should exists in the obj store
        logger.warning(e)
        return

    new_root = new_commit.root_id
    version = new_commit.get_version()

    if old_root == new_root:
        return

    old_root = old_root if old_root else ZERO_OBJ_ID

    differ = CommitDiffer(repo_id, version, old_root, new_root, False, False)

    return differ.diff()


def write_xls(sheet_name, head, data_list):
    """write listed data into excel
    """

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
    except Exception as e:
        logger.error(e)
        return None

    ws.title = sheet_name

    row_num = 0

    # write table head
    for col_num in range(len(head)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = head[col_num]

    # write table data
    for row in data_list:
        row_num += 1
        if row_num % 10000 == 0:
            time.sleep(0.5)
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    return wb


def utc_to_local(dt):
    # change from UTC timezone to current seahub timezone
    tz = pytz.timezone(TIME_ZONE)
    utc = dt.replace(tzinfo=datetime.timezone.utc)
    local = utc.astimezone(tz).replace(tzinfo=None)

    return local


def generate_file_audit_event_type(e):
    event_type_dict = {
        'file-download-web': ('web', ''),
        'file-download-share-link': ('share-link', ''),
        'file-download-api': ('API', e.device),
        'repo-download-sync': ('download-sync', e.device),
        'repo-upload-sync': ('upload-sync', e.device),
        'seadrive-download-file': ('seadrive-download', e.device),
    }

    if e.etype not in event_type_dict:
        event_type_dict[e.etype] = (e.etype, e.device if e.device else '')

    return event_type_dict[e.etype]


def export_event_log_to_excel(session, start_time, end_time, log_type, task_id):
    start_time = ast.literal_eval(start_time)
    end_time = ast.literal_eval(end_time)

    if not isinstance(start_time, (int, float)) or not isinstance(end_time, (int, float)):
        raise RuntimeError('Invalid time range parameter')

    if log_type not in ['fileaudit', 'fileupdate', 'permaudit', 'loginadmin']:
        raise RuntimeError('Invalid log_type parameter')

    with session() as session, CcnetDB() as ccnet_db, SeafileDB() as seafile_db:
        if log_type == 'fileaudit':
            stmt = select(FileAudit).where(FileAudit.timestamp.between(datetime.datetime.utcfromtimestamp(start_time),
                                                                       datetime.datetime.utcfromtimestamp(end_time)))
            stmt = stmt.order_by(desc(FileAudit.timestamp))
            res = session.scalars(stmt).all()

            head = ["User", "Type", "IP", "Device", "Date", "Library Name", "Library ID", "Library Owner", "File Path"]
            data_list = []
            repo_ids = set()
            for row in res:
                repo_ids.add(row.repo_id)
            repos = seafile_db.get_repo_info_by_ids(repo_ids)
            for ev in res:
                event_type, ev.show_device = generate_file_audit_event_type(ev)

                repo_id = ev.repo_id
                if repo_id in repos:
                    repo_name = repos[repo_id]['repo_name']
                    repo_owner = repos[repo_id]['owner']
                else:
                    repo_name = 'Deleted'
                    repo_owner = '--'

                username = ev.user if ev.user else 'Anonymous User'
                date = utc_to_local(ev.timestamp).strftime('%Y-%m-%d %H:%M:%S') if ev.timestamp else ''

                row = [username, event_type, ev.ip, ev.show_device,
                       date, repo_name, ev.repo_id, repo_owner, ev.file_path]
                data_list.append(row)

            wb = write_xls('file-access-logs', head, data_list)
            if not wb:
                raise RuntimeError('Failed to export file-access-logs to excel')

            target_dir = os.path.join('/tmp/seafile_events/', task_id)
            os.makedirs(target_dir, exist_ok=True)
            excel_name = 'file-access-logs.xlsx'
            target_path = os.path.join(target_dir, excel_name)
            wb.save(target_path)
        elif log_type == 'fileupdate':
            stmt = select(FileUpdate).where(FileUpdate.timestamp.between(datetime.datetime.utcfromtimestamp(start_time),
                                                                         datetime.datetime.utcfromtimestamp(end_time)))
            stmt = stmt.order_by(desc(FileUpdate.timestamp))
            res = session.scalars(stmt).all()

            head = ["User", "Date", "Library Name", "Library ID", "Library Owner", "Action"]
            data_list = []
            repo_ids = set()
            for row in res:
                repo_ids.add(row.repo_id)
            repos = seafile_db.get_repo_info_by_ids(repo_ids)

            for ev in res:
                repo_id = ev.repo_id
                if repo_id in repos:
                    repo_name = repos[repo_id]['repo_name']
                    repo_owner = repos[repo_id]['owner']
                else:
                    repo_name = 'Deleted'
                    repo_owner = '--'

                username = ev.user if ev.user else 'Anonymous User'
                date = utc_to_local(ev.timestamp).strftime('%Y-%m-%d %H:%M:%S') if ev.timestamp else ''
                row = [username, date, repo_name, ev.repo_id, repo_owner, ev.file_oper.strip()]
                data_list.append(row)

            wb = write_xls('file-update-logs', head, data_list)
            if not wb:
                raise RuntimeError('Failed to export file-update-logs to excel')

            target_dir = os.path.join('/tmp/seafile_events/', task_id)
            os.makedirs(target_dir, exist_ok=True)
            excel_name = 'file-update-logs.xlsx'
            target_path = os.path.join(target_dir, excel_name)
            wb.save(target_path)

        elif log_type == 'permaudit':
            stmt = select(PermAudit).where(PermAudit.timestamp.between(datetime.datetime.utcfromtimestamp(start_time),
                                                                       datetime.datetime.utcfromtimestamp(end_time)))
            stmt = stmt.order_by(desc(PermAudit.timestamp))
            res = session.scalars(stmt).all()

            head = ["From", "To", "Action", "Permission", "Library", "Folder Path", "Date"]
            data_list = []
            group_ids = set()
            repo_ids = set()
            for row in res:
                repo_ids.add(row.repo_id)
                if row.to.isdigit():
                    group_id = int(row.to)
                    group_ids.add(group_id)
            groups = ccnet_db.get_groups_by_ids(group_ids)
            repos = seafile_db.get_repo_info_by_ids(repo_ids)
            for ev in res:
                repo_id = ev.repo_id
                repo_name = repos[repo_id]['repo_name'] if repo_id in repos else 'Deleted'

                if '@' in ev.to:
                    to = ev.to
                elif ev.to.isdigit():
                    group_id = int(ev.to)
                    group = groups[group_id] if group_id in groups else None
                    to = group['group_name'] if group else 'Deleted'
                elif 'all' in ev.to:
                    to = 'Organization'
                else:
                    to = '--'

                if 'add' in ev.etype:
                    action = 'Add'
                elif 'modify' in ev.etype:
                    action = 'Modify'
                elif 'delete' in ev.etype:
                    action = 'Delete'
                else:
                    action = '--'

                if ev.permission == 'rw':
                    permission = 'Read-Write'
                elif ev.permission == 'r':
                    permission = 'Read-Only'
                else:
                    permission = '--'

                date = utc_to_local(ev.timestamp).strftime('%Y-%m-%d %H:%M:%S') if ev.timestamp else ''
                row = [ev.from_user, to, action, permission, repo_name, ev.file_path, date]
                data_list.append(row)

            wb = write_xls('perm-audit-logs', head, data_list)
            if not wb:
                raise RuntimeError('Failed to export perm-audit-logs to excel')

            target_dir = os.path.join('/tmp/seafile_events/', task_id)
            os.makedirs(target_dir, exist_ok=True)
            excel_name = 'perm-audit-logs.xlsx'
            target_path = os.path.join(target_dir, excel_name)
            wb.save(target_path)
        elif log_type == 'loginadmin':
            stmt = select(UserLogin).where(UserLogin.login_date.between(datetime.datetime.utcfromtimestamp(start_time),
                                                                        datetime.datetime.utcfromtimestamp(end_time)))
            stmt = stmt.order_by(desc(UserLogin.login_date))
            res = session.scalars(stmt).all()

            head = ["Name", "IP", "Status", "Time"]
            data_list = []
            for log in res:
                login_time = log.login_date.strftime("%Y-%m-%d %H:%M:%S")
                status = 'Success' if log.login_success else 'Failed'
                row = [log.username, log.login_ip, status, login_time]
                data_list.append(row)

            wb = write_xls('login-logs', head, data_list)
            if not wb:
                raise RuntimeError('Failed to export login-logs to excel')

            target_dir = os.path.join('/tmp/seafile_events/', task_id)
            os.makedirs(target_dir, exist_ok=True)
            excel_name = 'login-logs.xlsx'
            target_path = os.path.join(target_dir, excel_name)
            wb.save(target_path)


def export_org_event_log_to_excel(session, start_time, end_time, log_type, task_id, org_id):
    start_time = ast.literal_eval(start_time)
    end_time = ast.literal_eval(end_time)

    if not isinstance(start_time, (int, float)) or not isinstance(end_time, (int, float)):
        raise RuntimeError('Invalid time range parameter')

    if log_type not in ['fileupdate', 'permaudit', 'fileaudit']:
        raise RuntimeError('Invalid log_type parameter')

    with session() as session, CcnetDB() as ccnet_db, SeafileDB() as seafile_db:
        if log_type == 'fileupdate':
            stmt = select(FileUpdate).where(FileUpdate.timestamp.between(datetime.datetime.utcfromtimestamp(start_time),
                                                                         datetime.datetime.utcfromtimestamp(end_time)),
                                            FileUpdate.org_id == org_id)
            stmt = stmt.order_by(desc(FileUpdate.timestamp))
            res = session.scalars(stmt).all()

            head = ["User", "Date", "Library Name", "Library ID", "Library Owner", "Action"]
            data_list = []
            repo_ids = set()
            for row in res:
                repo_ids.add(row.repo_id)
            repos = seafile_db.get_repo_info_by_ids(repo_ids)

            for ev in res:
                repo_id = ev.repo_id
                if repo_id in repos:
                    repo_name = repos[repo_id]['repo_name']
                    repo_owner = repos[repo_id]['owner']
                else:
                    repo_name = 'Deleted'
                    repo_owner = '--'

                username = ev.user if ev.user else 'Anonymous User'
                date = utc_to_local(ev.timestamp).strftime('%Y-%m-%d %H:%M:%S') if ev.timestamp else ''
                row = [username, date, repo_name, ev.repo_id, repo_owner, ev.file_oper.strip()]
                data_list.append(row)

            wb = write_xls('file-update-logs', head, data_list)
            if not wb:
                raise RuntimeError('Failed to export file-update-logs to excel')

            target_dir = os.path.join('/tmp/seafile_events/', task_id)
            os.makedirs(target_dir, exist_ok=True)
            excel_name = 'file-update-logs.xlsx'
            target_path = os.path.join(target_dir, excel_name)
            wb.save(target_path)

        elif log_type == 'permaudit':
            stmt = select(PermAudit).where(PermAudit.timestamp.between(datetime.datetime.utcfromtimestamp(start_time),
                                                                       datetime.datetime.utcfromtimestamp(end_time)),
                                           PermAudit.org_id == org_id)
            stmt = stmt.order_by(desc(PermAudit.timestamp))
            res = session.scalars(stmt).all()
            head = ["From", "To", "Action", "Permission", "Library", "Folder Path", "Date"]
            data_list = []
            group_ids = set()
            repo_ids = set()
            for row in res:
                repo_ids.add(row.repo_id)
                if row.to.isdigit():
                    group_id = int(row.to)
                    group_ids.add(group_id)
            groups = ccnet_db.get_groups_by_ids(group_ids)
            repos = seafile_db.get_repo_info_by_ids(repo_ids)
            for ev in res:
                repo_id = ev.repo_id
                repo_name = repos[repo_id]['repo_name'] if repo_id in repos else 'Deleted'
                if '@' in ev.to:
                    to = ev.to
                elif ev.to.isdigit():
                    group_id = int(ev.to)
                    group = groups[group_id] if group_id in groups else None
                    to = group['group_name'] if group else 'Deleted'
                elif 'all' in ev.to:
                    to = 'Organization'
                else:
                    to = '--'

                if 'add' in ev.etype:
                    action = 'Add'
                elif 'modify' in ev.etype:
                    action = 'Modify'
                elif 'delete' in ev.etype:
                    action = 'Delete'
                else:
                    action = '--'

                if ev.permission == 'rw':
                    permission = 'Read-Write'
                elif ev.permission == 'r':
                    permission = 'Read-Only'
                else:
                    permission = '--'

                date = utc_to_local(ev.timestamp).strftime('%Y-%m-%d %H:%M:%S') if ev.timestamp else ''
                row = [ev.from_user, to, action, permission, repo_name, ev.file_path, date]
                data_list.append(row)

            wb = write_xls('perm-audit-logs', head, data_list)
            if not wb:
                raise RuntimeError('Failed to export perm-audit-logs to excel')

            target_dir = os.path.join('/tmp/seafile_events/', task_id)
            os.makedirs(target_dir, exist_ok=True)
            excel_name = 'perm-audit-logs.xlsx'
            target_path = os.path.join(target_dir, excel_name)
            wb.save(target_path)

        elif log_type == 'fileaudit':
            stmt = select(FileAudit).where(
                FileAudit.timestamp.between(datetime.datetime.utcfromtimestamp(start_time),
                                            datetime.datetime.utcfromtimestamp(end_time)),
                                            FileAudit.org_id == org_id)
            stmt = stmt.order_by(desc(FileAudit.timestamp))
            res = session.scalars(stmt).all()

            head = ["User", "Type", "IP", "Device", "Date", "Library Name", "Library ID", "Library Owner",
                    "File Path"]
            data_list = []
            repo_ids = set()
            for row in res:
                repo_ids.add(row.repo_id)
            repos = seafile_db.get_repo_info_by_ids(repo_ids)
            for ev in res:
                event_type, ev.show_device = generate_file_audit_event_type(ev)

                repo_id = ev.repo_id
                if repo_id in repos:
                    repo_name = repos[repo_id]['repo_name']
                    repo_owner = repos[repo_id]['owner']
                else:
                    repo_name = 'Deleted'
                    repo_owner = '--'

                username = ev.user if ev.user else 'Anonymous User'
                date = utc_to_local(ev.timestamp).strftime('%Y-%m-%d %H:%M:%S') if ev.timestamp else ''

                row = [username, event_type, ev.ip, ev.show_device,
                       date, repo_name, ev.repo_id, repo_owner, ev.file_path]
                data_list.append(row)

            wb = write_xls('file-access-logs', head, data_list)
            if not wb:
                raise RuntimeError('Failed to export file-access-logs to excel')

            target_dir = os.path.join('/tmp/seafile_events/', task_id)
            os.makedirs(target_dir, exist_ok=True)
            excel_name = 'file-access-logs.xlsx'
            target_path = os.path.join(target_dir, excel_name)
            wb.save(target_path)


def save_wiki_config(repo_id, username, wiki_config):
    dir_id = seafile_api.get_dir_id_by_path(repo_id, WIKI_CONFIG_PATH)
    if not dir_id:
        seafile_api.mkdir_with_parents(repo_id, '/', WIKI_CONFIG_PATH, username)
    tmp_content_path = posixpath.join(WIKI_FILE_TMP_DIR, WIKI_CONFIG_FILE_NAME)
    with open(tmp_content_path, 'wb') as f:
        f.write(wiki_config.encode())

    seafile_api.post_file(repo_id, tmp_content_path, WIKI_CONFIG_PATH, WIKI_CONFIG_FILE_NAME, username)
    os.remove(tmp_content_path)


def get_file_content_by_obj_id(repo_id, obj_id):
    if obj_id == ZERO_OBJ_ID:
        return ''
    f = fs_mgr.load_seafile(repo_id, 1, obj_id)
    b_content = f.get_content().decode()
    if not b_content.strip():
        return ''

    return b_content


def create_new_wiki_doc_by_old_wiki_path(page_name, file_parent_dir, new_repo_id, file_content, modifier,
                                         parent_path_to_page_id, pages, navigation, page_id_set, db_session_class,
                                         old_wiki_path_is_dir):

    sdoc_uuid = uuid.uuid4()
    sdoc_filename = page_name + '.sdoc'
    parent_dir = os.path.join(WIKI_PAGES_DIR, str(sdoc_uuid))
    sdoc_path = os.path.join(parent_dir, sdoc_filename)
    seafile_api.mkdir_with_parents(new_repo_id, '/', parent_dir.strip('/'), modifier)

    if old_wiki_path_is_dir:
        seafile_api.post_empty_file(new_repo_id, parent_dir, sdoc_filename, modifier)
    else:
        tmp_content_path = posixpath.join(WIKI_FILE_TMP_DIR, sdoc_filename)
        with open(tmp_content_path, 'wb') as f:
            f.write(file_content.encode())

        seafile_api.post_file(new_repo_id, tmp_content_path, parent_dir, sdoc_filename, modifier)
        os.remove(tmp_content_path)

    values = []
    file_is_dir = 0
    parent_path = normalize_path(parent_dir)
    repo_id_parent_path_md5 = md5_repo_id_parent_path(new_repo_id, parent_path)
    file_info = (sdoc_uuid.hex, new_repo_id, repo_id_parent_path_md5, parent_dir, sdoc_filename, file_is_dir)
    values.append(file_info)
    create_fileuuidmap_by_uuid(db_session_class, values)

    current_page_id = parent_path_to_page_id.get(file_parent_dir)
    new_page_id = gen_unique_id(page_id_set)
    page_id_set.add(new_page_id)

    if old_wiki_path_is_dir:
        tmp_path = posixpath.join(file_parent_dir, page_name)
        parent_path_to_page_id[tmp_path] = new_page_id

    gen_new_page_nav_by_id(navigation, new_page_id, current_page_id)
    new_page = {
        'id': new_page_id,
        'name': page_name,
        'path': sdoc_path,
        'icon': '',
        'docUuid': str(sdoc_uuid)
    }
    pages.append(new_page)


def is_sys_path(path):
    for sys_path in SYS_DIR_PATHS:
        if path.strip('/').startswith(sys_path):
            return True
    return False


def md5_repo_id_parent_path(repo_id, parent_path):
    parent_path = parent_path.rstrip('/') if parent_path != '/' else '/'
    return hashlib.md5((repo_id + parent_path).encode('utf-8')).hexdigest()


def normalize_path(path):
    return path.rstrip('/') if path != '/' else '/'


def create_fileuuidmap_by_uuid(db_session_class, values):
    with db_session_class() as session:
        sql = """INSERT INTO tags_fileuuidmap (uuid, repo_id, repo_id_parent_path_md5, parent_path, filename, is_dir)
                                     VALUES %s""" % ', '.join(
            ["('%s', '%s', '%s', '%s', '%s', %s)" % value for value in values])

        session.execute(text(sql))
        session.commit()


def generator_base64_code(length=4):
    possible = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789abcdefghijklmnopqrstuvwxyz0123456789'
    ids = random.sample(possible, length)
    return ''.join(ids)


def gen_unique_id(id_set, length=4):
    _id = generator_base64_code(length)

    while True:
        if _id not in id_set:
            return _id
        _id = generator_base64_code(length)


def gen_new_page_nav_by_id(navigation, page_id, current_id):
    new_nav = {
        'id': page_id,
        'type': 'page',
    }
    if current_id:
        for nav in navigation:
            if nav.get('type') == 'page' and nav.get('id') == current_id:
                sub_nav = nav.get('children', [])
                sub_nav.append(new_nav)
                nav['children'] = sub_nav
                return
            else:
                gen_new_page_nav_by_id(nav.get('children', []), page_id, current_id)
    else:
        navigation.append(new_nav)


def convert_wiki(old_repo_id, new_repo_id, username, db_session_class):
    new_commit_id = repo_data.get_repo_head_commit(old_repo_id)
    files = get_diff_files(old_repo_id, ZERO_OBJ_ID, new_commit_id)
    if not files:
        return
    added_files, deleted_files, added_dirs, deleted_dirs, modified_files, renamed_files, moved_files, \
    renamed_dirs, moved_dirs = files

    navigation = []
    pages = []
    page_id_set = set()
    parent_path_to_page_id = {}

    # first add home.md to new wiki page
    home_page_path = '/home.md'
    if new_commit_id is not None:
        file_id = seafile_api.get_file_id_by_commit_and_path(old_repo_id, new_commit_id, home_page_path)
    else:
        file_id = seafile_api.get_file_id_by_path(old_repo_id, home_page_path)

    if not file_id:
        home_page_content = ''
    else:
        home_page_content = get_file_content_by_obj_id(old_repo_id, file_id)
        if home_page_content:
            try:
                home_page_content = md2sdoc(home_page_content, username=username)
                home_page_content = json.dumps(home_page_content)
            except Exception as e:
                logger.warning('convert wiki from old_repo_id: %s to new_repo_id: %s, convert markdown: %s to sdoc failed, error: %s', old_repo_id, new_repo_id, home_page_path, e)
                home_page_content = ''

    parent_dir = os.path.dirname(home_page_path)
    base_name = os.path.basename(home_page_path)
    page_name = os.path.splitext(base_name)[0]
    create_new_wiki_doc_by_old_wiki_path(page_name, parent_dir, new_repo_id, home_page_content, username,
                                         parent_path_to_page_id, pages, navigation, page_id_set,
                                         db_session_class,
                                         False)

    for repo_dir in added_dirs:
        path = repo_dir.path
        if path == '/' or is_sys_path(path):
            continue

        dir_list = path.strip('/').split('/')
        pre_dir = '/'
        for dir1 in dir_list:
            parent_dir = pre_dir
            dir_page_name = dir1
            pre_dir = posixpath.join(pre_dir, dir1)
            if parent_path_to_page_id.get(pre_dir):
                continue

            file_content = ''
            modifier = username
            old_wiki_path_is_dir = True
            create_new_wiki_doc_by_old_wiki_path(dir_page_name, parent_dir, new_repo_id, file_content, modifier,
                                                 parent_path_to_page_id, pages, navigation, page_id_set,
                                                 db_session_class,
                                                 old_wiki_path_is_dir)

    for file in added_files:
        path = file.path.rstrip('/')
        obj_id = file.obj_id

        if is_sys_path(path):
            continue
        file_parent_dir = os.path.dirname(path)
        file_name = os.path.basename(path)
        modifier = file.modifier

        file_ext = os.path.splitext(file_name)[1][1:].lower()
        if path == '/home.md' or file_ext not in ('md', 'markdown'):
            continue

        markdown_content = get_file_content_by_obj_id(old_repo_id, obj_id)
        if markdown_content:
            try:
                file_content = md2sdoc(markdown_content, username=username)
                markdown_content = json.dumps(file_content)
            except Exception as e:
                logger.warning('convert wiki from old_repo_id: %s to new_repo_id: %s, convert markdown: %s to sdoc failed, error: %s', old_repo_id, new_repo_id, path, e)
                markdown_content = ''

        page_name = os.path.splitext(file_name)[0]
        create_new_wiki_doc_by_old_wiki_path(page_name, file_parent_dir, new_repo_id, markdown_content, modifier,
                                             parent_path_to_page_id, pages, navigation, page_id_set,
                                             db_session_class, False)

    wiki_config = {'version': 1, 'pages': pages, 'navigation': navigation}
    wiki_config = json.dumps(wiki_config)
    save_wiki_config(new_repo_id, username, wiki_config)
